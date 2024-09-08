import requests
import pandas as pd
import date_functions as datef
from datetime import timedelta, date
import datetime
import os
import dotenv
from openpyxl import load_workbook

dotenv.load_dotenv()


# CLIENT LIST

c_list = [
    "ajobrand",
    "dadri",
    "french",
    "infini",
    "mun",
    "othergirls",
    "talgui",
    "una",
]

c_list = ["french"]

directory = r"C:\Users\Samuel Kim\Documents\m&p"

# DICTS PARA SLICE DO CODIGO SKU

c_list_sku_0 = {"Cliente": ["mun", "infini"], "Index": 0}
c_list_sku_8 = {"Cliente": ["talgui", "dadri", "othergirls"], "Index": 8}
c_list_sku_9 = {"Cliente": ["ajobrand", "french", "una"], "Index": 9}

# INICIAL STOCK FOR EACH CLIENTE

est_inicial = {
    "ajobrand": "aJo Brand",
    "dadri": "Dadri",
    "french": "20231107",
    "infini": "Infini",
    "mun": "Mun",
    "nobu": "Nobu",
    "othergirls": "Other Girls",
    "paconcept": "P.A Concept",
    "talgui": "20231001",
    "una": "Una",
}


# API HEADER

payload = {}
files = {}

for client in c_list:
    headers = {
        "Authorization": os.getenv(f"ecco_aut_{client}"),
        "Content-Type": "application/json;charset=utf-8",
    }

    # In[1]: GET LIST OF PRODUCTS

    # ECCOSYS API (PRODUCTS)

    url_prod = "https://empresa.eccosys.com.br/api/produtos"
    response_prod = requests.request(
        "GET", url_prod, headers=headers, data=payload
    )

    dic_ecco_prod = response_prod.json()
    df_ecco_prod = pd.DataFrame(dic_ecco_prod)

    # ORGANIZE DF

    columns_to_keep = ["id", "codigo", "precoCusto", "precoDe", "preco"]
    df_ecco_prod = df_ecco_prod[columns_to_keep]

    columns_to_convert = ["id"]
    df_ecco_prod[columns_to_convert] = df_ecco_prod[columns_to_convert].astype(
        float
    )

    # In[2]: API: GET ALL REGISTERS OF STOCK MOVEMENT IN A CERTAIN DATE PERIOD

    # DATE FUCTIONS

    now = datetime.datetime.now()
    dtf = datef.dates(now)[2]

    # AUTOMATIC DATE
    from_date = pd.to_datetime(dtf).date() - timedelta(days=90)
    to_date = pd.to_datetime(dtf).date()

    # MANUAL DATE
    # to_date = date(2024, 1, 30) + timedelta(days=1)
    # from_date = to_date - timedelta(days=91)

    datatxt_from, dataname_from, datasql_from = datef.dates(from_date)
    datatxt_to, dataname_to, datasql_to = datef.dates(to_date)

    # ECCOSYS API (REGISTER OF STOCK MOVEMENT)

    url_reg = f"https://empresa.eccosys.com.br/api/estoques/registros?$offset=0&$count=1000000000&$fromDate={dataname_from}&$toDate={dataname_to}"

    response_ecco_prod_reg = requests.request(
        "GET", url_reg, headers=headers, data=payload, files=files
    )
    print(response_ecco_prod_reg)

    dic_ecco_prod_reg_all_new = response_ecco_prod_reg.json()
    df_ecco_prod_reg_all_new = pd.DataFrame(dic_ecco_prod_reg_all_new)

    # ORGANIZE DF

    df_ecco_prod_reg_all_new["data"] = pd.to_datetime(
        df_ecco_prod_reg_all_new["data"]
    )

    df_ecco_prod_reg_all_new["Cliente"] = client

    columns_to_keep = [
        "Cliente",
        "id",
        "data",
        "idProduto",
        "quantidade",
        "es",
    ]

    df_ecco_prod_reg_all_new = df_ecco_prod_reg_all_new[columns_to_keep]

    df_ecco_prod_reg_all_new.rename(columns={"id": "idMov"}, inplace=True)

    columns_to_convert = ["quantidade", "idMov", "idProduto"]
    df_ecco_prod_reg_all_new[columns_to_convert] = df_ecco_prod_reg_all_new[
        columns_to_convert
    ].astype(float)

    # JUST FOR NEW CLIENTS (CREATE FIRST REGISTER)

    if not os.path.exists(
        rf"C:\Users\Samuel Kim\Documents\m&p\registros\ecco_reg_stock_{client}.xlsx"
    ):
        df_ecco_prod_reg_all_new.to_excel(
            rf"C:\Users\Samuel Kim\Documents\m&p\registros\ecco_reg_stock_{client}.xlsx",
            index=False,
        )

    # In[3]: COMPARE AND ADD STOCK MOVEMENT REGISTERS

    # IMPORT LAST SAVED STOCK MOVEMENT REGISTER

    df_ecco_prod_reg_all_old = pd.read_excel(
        rf"C:\Users\Samuel Kim\Documents\m&p\registros\ecco_reg_stock_{client}.xlsx"
    )

    # COMPARE AND ADD NEW REGISTERS

    df_ecco_prod_reg_all_final = pd.merge(
        df_ecco_prod_reg_all_old,
        df_ecco_prod_reg_all_new,
        on=["idMov", "Cliente", "data", "es", "quantidade", "idProduto"],
        how="outer",
        indicator=True,
    )

    # DROP ROWS THAT WHERE EXCLUDED IN ECCOSYS

    df_index_to_remove = df_ecco_prod_reg_all_final[
        (df_ecco_prod_reg_all_final["data"].dt.date >= from_date)
        & (df_ecco_prod_reg_all_final["data"].dt.date < to_date)
        & (df_ecco_prod_reg_all_final["_merge"] == "left_only")
    ].index
    df_ecco_prod_reg_all_final.drop(df_index_to_remove, inplace=True)

    # ORGANIZE DF

    columns_to_keep = [
        "idMov",
        "Cliente",
        "data",
        "es",
        "quantidade",
        "idProduto",
    ]

    df_ecco_prod_reg_all_final = df_ecco_prod_reg_all_final[columns_to_keep]

    # In[4]: SAVE NEW REGISTER DF IN DB

    df_ecco_prod_reg_all_final.to_excel(
        rf"C:\Users\Samuel Kim\Documents\m&p\registros\ecco_reg_stock_{client}.xlsx",
        index=False,
    )

    # In[5]: CALCULATE FINAL DATE STOCK BALANCE

    # GROUP BY PRODUCT

    group_prod_reg_all_final = df_ecco_prod_reg_all_final.groupby("idProduto")

    # REMOVE ALL REGISTERS BEFORE 'es' = 'B'

    df_reg_tratado = pd.DataFrame()
    df_all_filter_b = pd.DataFrame()

    for produto, dados in group_prod_reg_all_final:
        filter_b = dados[dados["es"] == "B"]
        df_all_filter_b = pd.concat([df_all_filter_b, filter_b])

        if not filter_b.empty:
            max_b = filter_b["data"].max()
            after_b = dados[dados["data"] >= max_b]

            df_reg_tratado = pd.concat([df_reg_tratado, after_b])

        if filter_b.empty:
            df_reg_tratado = pd.concat([df_reg_tratado, dados])

    # SUM ALL REGISTERS BY PRODUCT

    df_reg_sum = (
        df_reg_tratado.groupby(["idProduto"])
        .agg({"quantidade": "sum"})
        .reset_index()
    )

    # BRING COLUMN 'codigo'

    df_reg_sum = pd.merge(
        df_reg_sum,
        df_ecco_prod,
        left_on="idProduto",
        right_on="id",
        how="left",
    )

    # ORGANIZE DF

    columns_to_keep = ["codigo", "quantidade"]
    df_reg_sum = df_reg_sum[columns_to_keep]

    df_reg_sum.rename(
        columns={"codigo": "Código", "quantidade": "Quantidade"}, inplace=True
    )

    # IMPORT INICIAL STOCK BEFORE REGISTERS DATE RANGE

    f_est_path = rf"C:\Users\Samuel Kim\Documents\m&p\estoque\estoque inicial\{client}_est_{est_inicial[client]}.xls"

    df_est_inicial = pd.read_excel(
        f_est_path, dtype={"Quantidade": float, "Código": str}
    )

    # ADD 'idProduct' ON 'df_est_inicial' SO I CAN USE IT TO REMOVE PRODUCT ROWS LATER

    df_est_inicial = pd.merge(
        df_est_inicial,
        df_ecco_prod[["codigo", "id"]],
        left_on="Código",
        right_on="codigo",
        how="left",
    )

    # STOCK FROM YESTERDAY SHOULD NOT MATTER FOR THE PRODUCTS THAT HAVE 'es' = 'B' TODAY

    df_filtered = pd.merge(
        df_est_inicial,
        df_all_filter_b,
        left_on="id",
        right_on="idProduto",
        how="left",
        indicator=True,
    )

    df_est_inicial = df_filtered[df_filtered["_merge"] == "left_only"].drop(
        columns=["_merge"]
    )

    # SUM INICIAL STOCK WITH ALL REGISTERS AND HAVE STOCK BALANCE

    df_est_final = pd.concat(
        [df_est_inicial, df_reg_sum], axis=0, ignore_index=True
    )

    df_est_final = (
        df_est_final.groupby("Código").agg({"Quantidade": "sum"}).reset_index()
    )

    # ADD DATE AND CLIENT COLUMN

    df_est_final["Data Saldo Estoque"] = pd.to_datetime(
        dataname_to
    ) - timedelta(days=1)

    df_est_final["Data Saldo Estoque"] = df_est_final[
        "Data Saldo Estoque"
    ].dt.date

    df_est_final["Cliente"] = client

    df_est_final = df_est_final[
        ["Cliente", "Data Saldo Estoque", "Código", "Quantidade"]
    ]

    # In[6]: SAVE STOCK DF IN DB

    # BRING STOCK DB

    df_stock_bd = pd.read_excel(
        rf"C:\Users\Samuel Kim\Documents\m&p\estoque\saldos\ecco_est_{client}.xlsx"
    )

    # IF I ALREDY HAVE THE df_est_final STOCK DATE, DONT APPEND

    if (
        df_est_final["Data Saldo Estoque"].unique()[0]
        not in df_stock_bd["Data Saldo Estoque"].dt.date.values
    ):
        book = load_workbook(
            rf"C:\Users\Samuel Kim\Documents\m&p\estoque\saldos\ecco_est_{client}.xlsx"
        )
        with pd.ExcelWriter(
            rf"C:\Users\Samuel Kim\Documents\m&p\estoque\saldos\ecco_est_{client}.xlsx",
            mode="a",
            if_sheet_exists="overlay",
            engine="openpyxl",
        ) as writer:
            df_est_final.to_excel(
                writer,
                index=False,
                header=False,
                sheet_name="Sheet1",
                startrow=writer.sheets["Sheet1"].max_row,
            )
